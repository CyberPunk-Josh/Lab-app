from PySide6.QtWidgets import QApplication, QMainWindow, QWidget, QMessageBox, QFileDialog
from PySide6.QtSql import QSqlDatabase, QSqlTableModel
from datetime import datetime
import pyqtgraph as pg
import openpyxl
import sys
import os

# Local files
from ui_subventana import Ui_Descripcion
from menu_design import *


class Subventana(QWidget, Ui_Descripcion):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.cancelarButton.clicked.connect(self.close)


class RModel(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()

        # Conection to database
        server = 'LAPTOP-2OCJA2M7\SQLEXPRESS'
        database = 'LabFluidos'
        driver = 'ODBC Driver 17 for SQL Server'

        db = QSqlDatabase.addDatabase('QODBC')
        db.setDatabaseName(f"Driver={driver};Server={server};Database={database};Trusted_Connection=yes;")
        if not db.open():
            print("Cannot connect to the database")
            sys.exit(True)

        # Desktop user path
        self.desktop = os.path.join(os.environ["HOMEPATH"], "Desktop")

        # creating the model
        self.model = QSqlTableModel()
        self.model.setTable("RegistroViscosimetro")
        self.model.select()

        # structure of objects for records
        self.records = [
            {"name": "Modelo 1", "valuesX": [], "valuesY": [], "color": "green", "symbol": "o", "VA": 0, "VP": 0,
             "PC": 0, "Temperatura": "23 [ºC]", "tethaValues": []},
            {"name": "Modelo 2", "valuesX": [], "valuesY": [], "color": "blue", "symbol": "o", "VA": 0, "VP": 0,
             "PC": 0, "Temperatura": "60 [ºC]", "tethaValues": []},
            {"name": "Modelo 3", "valuesX": [], "valuesY": [], "color": "pink", "symbol": "o", "VA": 0, "VP": 0,
             "PC": 0, "Temperatura": "65 [ºC]", "tethaValues": []},
            {"name": "Modelo 4", "valuesX": [], "valuesY": [], "color": "red", "symbol": "o", "VA": 0, "VP": 0,
             "PC": 0, "Temperatura": "70 [ºC]", "tethaValues": []},
            {"name": "Modelo 5", "valuesX": [], "valuesY": [], "color": "brown", "symbol": "o", "VA": 0, "VP": 0,
             "PC": 0, "Temperatura": "82 [ºC]", "tethaValues": []}
        ]

        # Adding object's name to comboBox
        for record in self.records:
            self.RM_ComboBox.addItem(record["name"] + " " + record["Temperatura"])

        # Build graphic for reological model
        self.buildGraphics()

        # Sub window for mesage
        self.subventana = Subventana()

        # Signals
        self.RM_Graph.clicked.connect(self.getReologicalValues)
        self.RM_Clear.clicked.connect(self.clearReologicalValues)
        self.RM_ComboBox.currentTextChanged.connect(self.comboChanged)
        self.RM_Save.clicked.connect(self.showSaveWindow)
        self.RM_Rep.clicked.connect(self.getReologicalReport)
        self.subventana.aceptarButton.clicked.connect(self.saveValuesToDB)



    def buildGraphics(self):
        self.RM_Widget.addLegend()
        self.RM_Widget.setBackground("w")
        self.graphics = []
        for record in self.records:
            plot = self.RM_Widget.plot(record["valuesX"], record["valuesY"], name=record["Temperatura"],
                                       pen=pg.mkPen(record["color"], width=3), symbol=record["symbol"])
            self.graphics.append(plot)
        self.RM_Widget.showGrid(x=True, y=True)
        self.RM_Widget.setTitle("Modelo Reològico", sizer="24px")
        styles = {"color": "#000", "font-size": "20px"}
        self.RM_Widget.setLabel("left", "t [Dinas/cm2]", **styles)
        self.RM_Widget.setLabel("bottom", "y [seg-1]", **styles)

    def getReologicalValues(self):
        try:
            # Getting values from table in UI
            valuesInX = [
                int(self.RM_TableWidget.item(0, 0).text()),
                int(self.RM_TableWidget.item(1, 0).text()),
                int(self.RM_TableWidget.item(2, 0).text()),
                int(self.RM_TableWidget.item(3, 0).text()),
                int(self.RM_TableWidget.item(4, 0).text()),
                int(self.RM_TableWidget.item(5, 0).text())
            ]

            valuesInY = [
                float(self.RM_TableWidget.item(0, 1).text()),
                float(self.RM_TableWidget.item(1, 1).text()),
                float(self.RM_TableWidget.item(2, 1).text()),
                float(self.RM_TableWidget.item(3, 1).text()),
                float(self.RM_TableWidget.item(4, 1).text()),
                float(self.RM_TableWidget.item(5, 1).text())
            ]

            index = self.RM_ComboBox.currentIndex()

            for value in valuesInX:
                self.records[index]["valuesX"].append(round(value * 1.703, 3))
            for value in valuesInY:
                self.records[index]["tethaValues"].append(value)
                self.records[index]["valuesY"].append(round(value * 5.1109, 3))

            # CALCULATING REOLOGICAL VALUES
            VA = round(float(self.RM_TableWidget.item(5, 1).text()) / 2, 2)
            VP = round(float(self.RM_TableWidget.item(5, 1).text()) - float(self.RM_TableWidget.item(4, 1).text()), 2)
            PC = round(float(self.RM_TableWidget.item(4, 1).text()) - VP, 2)

            # ADDING REOLOGICAL DATA TO OBJECT
            self.records[index]["VA"] = VA
            self.records[index]["VP"] = VP
            self.records[index]["PC"] = PC

            # ADDING REOLOGICAL DATA TO INPUTS IN UI
            self.RM_Va.setValue(self.records[index]["VA"])
            self.RM_Vp.setValue(self.records[index]["VP"])
            self.RM_Pc.setValue(self.records[index]["PC"])

            # adding values into graph and updating graphics
            self.graphics[index].setData(self.records[index]["valuesX"], self.records[index]["valuesY"])
        except:
            QMessageBox.information(self, "Error", "Todos los campos son obligatorios, revisa tu informarción")

    # Clear reological values and graphics
    def clearReologicalValues(self):
        # clear Values
        self.RM_TableWidget.item(0, 1).setText("")
        self.RM_TableWidget.item(1, 1).setText("")
        self.RM_TableWidget.item(2, 1).setText("")
        self.RM_TableWidget.item(3, 1).setText("")
        self.RM_TableWidget.item(4, 1).setText("")
        self.RM_TableWidget.item(5, 1).setText("")
        self.RM_Va.setValue(0)
        self.RM_Vp.setValue(0)
        self.RM_Pc.setValue(0)

        index = self.RM_ComboBox.currentIndex()

        # clear values from list
        self.records[index]["valuesX"].clear()
        self.records[index]["valuesY"].clear()
        self.records[index]["tethaValues"].clear()
        self.records[index]["VA"] = 0
        self.records[index]["VP"] = 0
        self.records[index]["PC"] = 0

        # updating graphics
        self.graphics[index].setData(self.records[index]["valuesX"], self.records[index]["valuesY"])

    # Recharging of values when user changes index in comboBox
    def comboChanged(self):
        index = self.RM_ComboBox.currentIndex()
        if len(self.records[index]["tethaValues"]) > 0:
            self.RM_TableWidget.item(0, 1).setText(str(self.records[index]["tethaValues"][0]))
            self.RM_TableWidget.item(1, 1).setText(str(self.records[index]["tethaValues"][1]))
            self.RM_TableWidget.item(2, 1).setText(str(self.records[index]["tethaValues"][2]))
            self.RM_TableWidget.item(3, 1).setText(str(self.records[index]["tethaValues"][3]))
            self.RM_TableWidget.item(4, 1).setText(str(self.records[index]["tethaValues"][4]))
            self.RM_TableWidget.item(5, 1).setText(str(self.records[index]["tethaValues"][5]))
            self.RM_Va.setValue(self.records[index]["VA"])
            self.RM_Vp.setValue(self.records[index]["VP"])
            self.RM_Pc.setValue(self.records[index]["PC"])
        else:
            self.RM_TableWidget.item(0, 1).setText("")
            self.RM_TableWidget.item(1, 1).setText("")
            self.RM_TableWidget.item(2, 1).setText("")
            self.RM_TableWidget.item(3, 1).setText("")
            self.RM_TableWidget.item(4, 1).setText("")
            self.RM_TableWidget.item(5, 1).setText("")
            self.RM_Va.setValue(0)
            self.RM_Vp.setValue(0)
            self.RM_Pc.setValue(0)

    def showSaveWindow(self):
        # showing sub window to add description
        self.subventana.show()

    # TODO Add a field validator before saving values to database
    def saveValuesToDB(self):
        try:
            # getting tetha values
            valuesInY = [
                int(float(self.RM_TableWidget.item(0, 1).text())),
                int(float(self.RM_TableWidget.item(1, 1).text())),
                int(float(self.RM_TableWidget.item(2, 1).text())),
                int(float(self.RM_TableWidget.item(3, 1).text())),
                int(float(self.RM_TableWidget.item(4, 1).text())),
                int(float(self.RM_TableWidget.item(5, 1).text()))
            ]
            # getting description text
            description = self.subventana.textEdit.toPlainText()
            time = datetime.now()
            timeSQL = time.strftime("%Y-%m-%d %H:%M:%S")
            # insert data into sql server
            new_row = self.model.rowCount()
            self.model.insertRow(new_row)
            self.model.setData(self.model.index(new_row, 1), valuesInY[0])
            self.model.setData(self.model.index(new_row, 2), valuesInY[1])
            self.model.setData(self.model.index(new_row, 3), valuesInY[2])
            self.model.setData(self.model.index(new_row, 4), valuesInY[3])
            self.model.setData(self.model.index(new_row, 5), valuesInY[4])
            self.model.setData(self.model.index(new_row, 6), valuesInY[5])
            self.model.setData(self.model.index(new_row, 7), description)
            self.model.setData(self.model.index(new_row, 8), timeSQL)
            self.model.submit()
            self.model.select()
            QMessageBox.information(self, "Exito", "Se han guardado los datos exitosamente")
            self.subventana.close()
            self.subventana.textEdit.setText("")
        except:
            QMessageBox.information(self, "Error", "Todos los campos son obligatorios, revisa tu información")
            self.subventana.close()
            self.subventana.textEdit.setText("")

    def getReologicalReport(self):
        # open window to save file
        userPath = QFileDialog.getExistingDirectory(
            parent=self,
            caption='Selecciona una ubicacion',
            dir=self.desktop
        )
        # create excel file
        wb = openpyxl.Workbook()
        ws = wb.active

        # get values of the active index
        index = self.RM_ComboBox.currentIndex()
        if len(self.records[index]["tethaValues"]) > 0:
            # structuring data
            rows = [
                ['y', 't'],
                [self.records[index]["valuesX"][0], self.records[index]["valuesY"][0]],
                [self.records[index]["valuesX"][1], self.records[index]["valuesY"][1]],
                [self.records[index]["valuesX"][2], self.records[index]["valuesY"][2]],
                [self.records[index]["valuesX"][3], self.records[index]["valuesY"][3]],
                [self.records[index]["valuesX"][4], self.records[index]["valuesY"][4]],
                [self.records[index]["valuesX"][5], self.records[index]["valuesY"][5]],
            ]
            # inserting data into excel file
            for row in rows:
                ws.append(row)

            # creating chart
            c1 = openpyxl.chart.ScatterChart()
            c1.title = self.records[index]["name"] + '- ' + self.records[index]["Temperatura"]
            c1.style = 2
            c1.y_axis.title = 't'
            c1.x_axis.title = 'y'

            xValues = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=7)
            yValues = openpyxl.chart.Reference(ws, min_col=2, min_row=2, max_row=7)

            series = openpyxl.chart.Series(yValues, xValues, title='Modelo Reològico')

            c1.append(series)

            # style line
            s1 = c1.series[0]
            s1.marker = openpyxl.chart.marker.Marker('x')  # Add marker for intersection
            s1.marker.graphicalProperties.solidFill = '000000'  # Color for marker
            s1.graphicalProperties.line.width = 30000
            s1.smooth = True  # Make the line smooth

            # adding reological data
            ws['E2'] = 'VA'
            ws['F2'] = self.records[index]["VA"]
            ws['G2'] = '[CP]'

            ws['E3'] = 'VP'
            ws['F3'] = self.records[index]["VP"]
            ws['G3'] = '[CP]'

            ws['E4'] = 'PC'
            ws['F4'] = self.records[index]["PC"]
            ws['G4'] = '[ADIM]'
            # adding chat and saving file
            ws.add_chart(c1, 'B10')
            wb.save(f"{userPath}/report.xlsx")
            # success message for user
            QMessageBox.information(self, "Correcto", "Reporte guardado exitosamente")

        else:
            QMessageBox.information(self, "Error", "Todos los campos son obligatorios, revisa tu información")
